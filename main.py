import os
from datetime import datetime, timedelta

import customtkinter as ctk
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Functions
def validate_numeric_input(new_value):
    if new_value == "":
        return True
    try:
        float(new_value)
        return True
    except ValueError:
        return False


def on_button_click():
    item_name = item_name_entry_box.get()
    item_price = item_price_entry_box.get()
    selected_date = date_combobox.get()

    if item_name and item_price and selected_date != "Select Date":
        workbook = load_workbook(file_path)
        sheet = workbook.active
        next_row = sheet.max_row + 1
        sheet[f"A{next_row}"] = item_name
        sheet[f"B{next_row}"] = item_price
        sheet[f"C{next_row}"] = selected_date
        workbook.save(file_path)
        message_label.configure(text="Information noted!")
        main_window.after(2000, lambda: message_label.configure(text=""))
    else:
        message_label.configure(text="Please fill all fields!")
        main_window.after(2000, lambda: message_label.configure(text=""))


def day_suffix(day):
    if 4 <= day <= 20 or 24 <= day <= 30:
        return "th"
    else:
        suffixes = {1: "st", 2: "nd", 3: "rd"}
        return suffixes.get(day % 10, "th")


# Database stuff
folder_path = "C:/Expense-Tracker"
file_path = f"{folder_path}/Expenses.xlsx"
workbook = Workbook()
if os.path.exists(file_path):
    workbook = load_workbook(file_path)
else:
    if os.path.exists(folder_path):
        os.makedirs(file_path)
    else:
        os.makedirs(folder_path)
        open(file_path, "w").close()

workbook.save(file_path)

sheet = workbook.active
sheet["A1"] = "Item"
sheet["B1"] = "Price in ₱"
sheet["C1"] = "Date of purchase"

column_widths = [20, 20, 20]
column_range = range(1, 4)
row_range = range(1, sheet.max_row + 1)
for col_num, width in enumerate(column_widths, start=1):
    column_letter = get_column_letter(col_num)
    sheet.column_dimensions[column_letter].width = width

workbook.save(file_path)

# Main window stuff 
main_window = ctk.CTk()
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")
main_window.title("Tkinter GUI")
screen_width = main_window.winfo_screenwidth()
screen_height = main_window.winfo_screenheight()
window_width = 500
window_height = 330
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)
main_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
main_window.resizable(False, False)

# Widget stuff
main_frame = ctk.CTkFrame(master=main_window)
main_frame.pack(pady=15, padx=20, fill="both", expand=True)
frame_label = ctk.CTkLabel(master=main_frame, text="Expense Tracker", font=("Roboto", 15))
frame_label.pack(pady=10, padx=13)

# Product info stuff
item_frame = ctk.CTkFrame(master=main_frame)
item_frame.pack(pady=10, padx=13)
item_frame.grid_columnconfigure(0, weight=1)
item_frame.grid_columnconfigure(1, weight=1)
item_frame.grid_columnconfigure(2, weight=1)

item_name_entry_box = ctk.CTkEntry(master=item_frame, placeholder_text="Item", placeholder_text_color="white")
item_name_entry_box.grid(row=0, column=0, padx=13, pady=10, sticky="nsew")

item_price_entry_box = ctk.CTkEntry(master=item_frame, placeholder_text="Price", placeholder_text_color="white")
item_price_entry_box.grid(row=0, column=1, padx=13, pady=10, sticky="nsew")
validate_cmd = (main_window.register(validate_numeric_input), '%P')
item_price_entry_box.configure(validate='key', validatecommand=validate_cmd)

start_date = datetime.now().date()
end_date = start_date.replace(month=12, day=31)
delta = timedelta(days=1)
date_list = []
current_date = start_date
while current_date <= end_date:
    date_list.append(current_date.strftime("%B %d"))
    current_date += delta

date_combobox = ctk.CTkComboBox(master=item_frame, values=date_list, state="readonly")
date_combobox.grid(row=0, column=2, padx=13, pady=10, sticky="nsew")
date_combobox.set("Select Date")

# Enter information 
enter_button = ctk.CTkButton(master=main_frame, text="Enter", command=on_button_click)
enter_button.pack(pady=5, padx=13)
message_label = ctk.CTkLabel(master=main_frame, text="")
message_label.pack(pady=3, padx=4)

# Monthly report 
report_label = ctk.CTkLabel(master=main_frame, text="Monthly report:")
report_label.pack(pady=5, padx=13, anchor="w")

report_frame = ctk.CTkScrollableFrame(master=main_frame)
report_frame.pack(pady=8, padx=13, fill="both", expand=True)
last_day_of_month = (start_date.replace(day=1, month=start_date.month + 1) - timedelta(days=1)).day
text_to_display = ""

report_workbook = load_workbook(file_path)
report_sheet = report_workbook.active

item_count = report_sheet["A"]
number_of_items = sum(1 for cell in item_count if cell.value is not None) - 1

item_price_range = report_sheet["B"][1:]
total_price = sum(float(cell.value) for cell in item_price_range if cell.value is not None)

if start_date.day == last_day_of_month:
    formatted_date = start_date.strftime("%d{suffix} of %B, %Y").format(suffix=day_suffix(start_date.day))
    text_to_display = f"By the end of the month ({formatted_date}), you bought {number_of_items} items and spent ₱{total_price} within this month."
else:
    text_to_display = "It's not the end of the month yet. No need for a report."

text_label = ctk.CTkLabel(master=report_frame, text=text_to_display, wraplength=340, justify="center")
text_label.pack(pady=10, padx=13)

main_window.mainloop()
